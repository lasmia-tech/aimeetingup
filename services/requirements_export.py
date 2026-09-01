"""추출된 요구사항 목록을 요구사항정의서 엑셀 양식에 채워 넣는 서비스.

템플릿(templates/요구사항정의서_템플릿.xlsx)은 제목 아래에 표만 있는 단순한
구조라, requirements_field_map.json의 table.columns에 정의된 필드만큼만
값이 채워진다(정의에 없는 item의 키는 무시된다). rows(기본 30줄)만큼은 표
서식(테두리 등)이 미리 입혀져 있으므로 그 범위 안에서는 값만 채우면 되고,
넘치는 경우에만 마지막 서식 행을 복사해 행을 늘린다.
"""

import copy
import json
from pathlib import Path

import openpyxl


def _load_field_map(field_map_path: str) -> dict:
    with open(field_map_path, encoding="utf-8") as f:
        return json.load(f)


def fill_requirements_template(
    items: list[dict],
    template_path: str,
    field_map_path: str,
    output_path: str,
) -> str:
    """요구사항 목록을 실제 엑셀 양식 파일의 표에 채워 넣는다.

    Args:
        items: extract_requirements()가 반환한(또는 화면에서 수정한) 요구사항
            목록. 각 항목은 field_map의 table.columns 키(no 제외)를 가진다.
        template_path: 원본 요구사항정의서 양식(.xlsx) 파일 경로.
        field_map_path: 표 위치/컬럼 매핑이 담긴 JSON 파일 경로.
        output_path: 값이 채워진 결과 파일을 저장할 경로.

    Returns:
        저장된 결과 파일의 경로(output_path와 동일한 값).
    """
    field_map = _load_field_map(field_map_path)
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    table = field_map["table"]
    start_row = table["start_row"]
    formatted_rows = table["rows"]
    columns = table["columns"]

    # 미리 서식이 입혀진 행 수(formatted_rows)보다 항목이 많으면, 마지막 서식
    # 행의 스타일을 복사해서 필요한 만큼 행을 추가한다.
    if len(items) > formatted_rows:
        extra = len(items) - formatted_rows
        last_formatted_row = start_row + formatted_rows - 1
        ws.insert_rows(last_formatted_row + 1, amount=extra)
        for col_letter in columns.values():
            src_cell = ws[f"{col_letter}{last_formatted_row}"]
            for r in range(last_formatted_row + 1, last_formatted_row + 1 + extra):
                dst_cell = ws[f"{col_letter}{r}"]
                dst_cell.font = copy.copy(src_cell.font)
                dst_cell.border = copy.copy(src_cell.border)
                dst_cell.alignment = copy.copy(src_cell.alignment)
                dst_cell.fill = copy.copy(src_cell.fill)

    for i, item in enumerate(items):
        row = start_row + i
        ws[f"{columns['no']}{row}"] = i + 1
        for field, col_letter in columns.items():
            if field == "no":
                continue
            ws[f"{col_letter}{row}"] = item.get(field, "")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
