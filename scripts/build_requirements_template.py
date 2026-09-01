"""요구사항정의서 엑셀 양식(templates/요구사항정의서_템플릿.xlsx)을 생성하는 스크립트.

이 스크립트는 실제 템플릿 파일의 최신 레이아웃(2단 헤더 + 4행부터 시작하는
데이터 표)을 코드로 재현해서, 필요할 때 다시 만들어낼 수 있도록 유지하는
용도다. 서식을 조정하고 싶으면 이 스크립트를 고쳐서 다시 실행하면 된다
(python scripts/build_requirements_template.py).
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUTPUT_PATH = Path(__file__).resolve().parent.parent / "templates" / "요구사항정의서_템플릿.xlsx"

HEADER_ROW = 2  # 대부분의 헤더 셀이 2~3행에 걸쳐 병합된다.
SUB_HEADER_ROW = 3  # "업무구분"의 대/중/소분류만 3행에 별도로 존재한다.
TABLE_START_ROW = 4
TABLE_ROWS = 30  # 미리 서식을 입혀둘 데이터 행 수. 초과분은 코드에서 서식을 복사해 행을 추가한다.

# (열 문자, 헤더 텍스트, 열 너비). "업무구분"(F~H)만 예외적으로 3행에 하위 헤더가 있다.
COLUMNS = [
    ("A", "No", 6),
    ("B", "요구사항ID", 10),
    ("C", "요구사항명", 16),
    ("D", "요구사항설명", 13),
    ("E", "해결방안/고려사항", 18),
    ("F", "업무구분", 7.4),
    ("G", "업무구분", 7.4),
    ("H", "업무구분", 7.4),
    ("I", "구분", 5.5),
    ("J", "요구자", 7.4),
    ("K", "요건원천", 12.2),
    ("L", "수용여부", 9.2),
    ("M", "우선순위", 9.2),
    ("N", "비고", 5.5),
]
SUB_HEADERS = {"F": "대분류", "G": "중분류", "H": "소분류"}
# 업무구분(F~H)은 대/중/소분류로 나뉘어 하위 헤더가 있어서 세로 병합에서 제외한다.
NO_VERTICAL_MERGE = {"F", "G", "H"}

thin = Side(style="thin", color="999999")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def build() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "요구사항정의서"

    for col, _label, width in COLUMNS:
        ws.column_dimensions[col].width = width

    # 제목
    last_col = COLUMNS[-1][0]
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "요구사항정의서"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # 표 헤더 (업무구분 3칸만 가로 병합 + 대/중/소분류 하위 헤더, 나머지는 2~3행 세로 병합)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    ws.merge_cells(f"F{HEADER_ROW}:H{HEADER_ROW}")
    for col, label, _width in COLUMNS:
        cell = ws[f"{col}{HEADER_ROW}"]
        cell.value = label
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        if col not in NO_VERTICAL_MERGE:
            ws.merge_cells(f"{col}{HEADER_ROW}:{col}{SUB_HEADER_ROW}")
        ws[f"{col}{SUB_HEADER_ROW}"].border = border

    for col, sub_label in SUB_HEADERS.items():
        cell = ws[f"{col}{SUB_HEADER_ROW}"]
        cell.value = sub_label
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터 행 서식만 미리 입혀둔다 (값은 비워둠)
    for r in range(TABLE_START_ROW, TABLE_START_ROW + TABLE_ROWS):
        for col, _label, _width in COLUMNS:
            cell = ws[f"{col}{r}"]
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in ("D", "E")))

    # 요건원천은 이 앱이 항상 회의록에서만 추출하므로 첫 데이터 행에 예시로 남겨둔다.
    ws[f"K{TABLE_START_ROW}"] = "회의록"

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"저장됨: {OUTPUT_PATH}")


if __name__ == "__main__":
    build()
